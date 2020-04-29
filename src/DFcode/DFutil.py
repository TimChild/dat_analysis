import pandas as pd
import numpy as np
from functools import wraps
from src.Configs import Main_Config as cfg
from src.CoreUtil import verbose_message
import src.CoreUtil as CU
import pickle
import os
import openpyxl
from typing import Tuple, List, MutableMapping, Union


def get_excel(path, index_col: Union[list, int] = 0, header: Union[list, int] = 0, dtype=None) -> pd.DataFrame:
    """Returns excel file as pd.DataFrame, handles when excel is already open"""
    na_values = ['-NaN', 'NaN', '-nan', 'nan']
    while True:
        try:
            excel = pd.read_excel(path, index_col=index_col, header=header, dtype=dtype, na_values=na_values,
                                  keep_default_na=False)
            return excel
        except PermissionError:
            print('PermissionError: Please close file in excel then press any key to continue')
            input()


def getexceldf(path, comparisondf=None, dtypes: dict = None, name: str = None) -> pd.DataFrame:
    """Returns excel df at given path, or will ask for user input comparison df provided"""
    if name is None:
        name = f'Generic Comparison'
    if os.path.isfile(path):
        index_columns, headers = _get_excel_header_index(path)
        exceldf = get_excel(path, index_col=index_columns, header=headers, dtype=dtypes)
        if comparisondf is not None:
            df = compare_pickle_excel(comparisondf, exceldf,
                                      name)  # Returns either pickle or excel df depending on input
        else:
            df = exceldf
    elif comparisondf is not None:
        df = comparisondf
    else:
        raise ValueError(f'No dataframe to return, none given as "dfcompare" param and none at [{path}]')
    return df


def _get_excel_header_index(xlsx_path) -> Tuple[List[int], List[int]]:
    """Reads top left of excel sheet to figure out how many headers and index's
    (assumes column levels are labelled if there is more than one)"""
    book = openpyxl.load_workbook(xlsx_path)
    sheet = book.active
    if sheet['A1'].value is not None:
        return 0, 0  # No column depth, and no way to tell what the index depth is
    col_depth = 0
    index_depth = 0
    while col_depth < 10:
        col_depth += 1
        if sheet['A' + str(col_depth + 1)].value is not None:
            # Hit the index titles so no more column depth
            break
    while index_depth < 10:
        index_depth += 1
        if sheet.cell(row=index_depth + 1, column=1).value is not None and sheet.cell(row=index_depth + 1,
                                                                                      column=1).value[:5] != 'level':
            # Hit the first column heading
            break
    return [0, index_depth - 1], [0, col_depth - 1]  # -1 so e.g. header depth of 2 indexes (0,1)


def open_xlsx(filepath):
    def _is_excel(excelfilepath):
        if excelfilepath[-4:] == 'xlsx':
            return True
        else:
            raise TypeError(f'Filepath points to a non-excel file at "{excelfilepath}"')

    if _is_excel(filepath):
        if os.path.isfile(filepath):
            os.startfile(filepath)
        else:
            raise FileNotFoundError(f'No file found at "{filepath}"')
    return None


def change_in_excel(path) -> pd.DataFrame:
    """Lets user change df in excel. Does not save changes by default!!!"""
    open_xlsx(path)
    input(f'After finished editing and changes are saved press any key to continue')
    df = getexceldf(path)
    # region Verbose SetupDF change_in_excel
    if cfg.verbose is True:
        verbose_message('DF loaded from excel. Not saved to pickle by default!!!')
    # endregion
    return df


def protect_data_from_reindex(func):
    if getattr(func, '_decorated', False):  # Prevent double wrapping of function
        return func

    @wraps(func)
    def wrapper(*args, **kwargs):
        df = args[0]  # type: pd.DataFrame
        assert isinstance(df, pd.DataFrame)
        if df.index.names[0] is not None:
            df.reset_index(inplace=True)
            print(f'WARNING: You were just saved from deleting the data [{df.index.names}] while reassigning index.'
                  f'\nYou might need to keep track of your index more closely')
        ret = func(*args, **kwargs)
        return ret

    wrapper._decorated = True
    return wrapper


def compare_pickle_excel(dfpickle, dfexcel, name='...[NAME]...') -> pd.DataFrame:
    df = dfpickle
    if _compare_to_df(dfpickle, dfexcel) is False:
        ans = CU.option_input(f'DFs for {name} have a different pickle and excel version of DF '
                              f'do you want to use excel version?', {'yes': True, 'no': False})
        if ans is True:  # Replace pickledf with exceldf
            df = dfexcel
    return df


def _compare_to_df(dfone, dftwo):
    """Returns true if equal, False if not.. Takes advantage of less precise comparison
    of assert_frame_equal in pandas"""
    try:
        pd.testing.assert_frame_equal(dfone, dftwo, check_dtype=False)
        return True
    except AssertionError as e:
        # region Verbose DatDF compare_to_df
        if cfg.verbose is True:
            verbose_message(f'Verbose[][_compare_to_df] - Difference between dataframes is [{e}]')
        # endregion
        return False


def load_from_pickle(path, cls):
    """Returns instance of class at path if it exists, otherwise returns None"""
    if os.path.isfile(path):
        with open(path, 'rb') as f:
            inst = pickle.load(f)  # inspect.stack()[1][3] == __new__ required to stop loop here
        inst.loaded = True
        if not isinstance(inst, cls):  # Check if loaded version is the right class
            raise TypeError(f'File saved at {path} is not of the type {cls}')
        return inst
    else:
        return None


def temp_reset_index(func):
    """Decorator that temporarily resets index then returns it to what it was before. Requires either df as first
     argument or an object a df at obj.df as first argument. Mostly to be used on SetupDF or DatDF methods"""
    _flag = None

    def _getdfindexnames(*args):
        nonlocal _flag
        if isinstance(args[0], pd.DataFrame):
            df = args[0]
            _flag = 0
        elif isinstance(args[0].df, pd.DataFrame):
            df = args[0].df
            _flag = 1
        else:
            raise ValueError('temp_reset_index requires df or object with .df as first arg')
        indexnames = df.index.names
        return indexnames

    def _setdfindex(*args, indexnames: list = (None)):
        nonlocal _flag
        if indexnames[0] is not None:
            if _flag == 0:
                _resetdfindex(*args)
                args[0].set_index(indexnames, inplace=True)
            elif _flag == 1:
                _resetdfindex(*args)
                args[0].df.set_index(indexnames, inplace=True)
        return None

    def _resetdfindex(*args):
        nonlocal _flag
        if _flag == 0:
            if args[0].index.names[0] is not None:
                args[0].reset_index(inplace=True)
        elif _flag == 1:
            if args[0].df.index.names[0] is not None:
                args[0].df.reset_index(inplace=True)
        return None

    @wraps(func)
    def wrapper(*args, **kwargs):
        indexnames = _getdfindexnames(*args)
        _resetdfindex(*args)
        ret = func(*args, **kwargs)
        _setdfindex(*args, indexnames=indexnames)
        return ret

    return wrapper


def add_col_label(df, new_col, on_cols, level=1):  # 2/2/20 I think it's something in here that requires the '.' address
    """Add a new label to an existing column (e.g. add a second level)"""

    def _new_level_emptycols(df, level=1, address='top'):
        if level == 1:
            return dict(zip(df.columns, np.repeat('.', df.shape[1])))
        else:
            if address == 'full':
                return dict(zip([x for x in df.columns], np.repeat('.', df.shape[1])))
            elif address == 'top':
                return dict(zip([x[0] for x in df.columns], np.repeat('.', df.shape[1])))
            else:
                raise ValueError(f'Address "{address}" is not valid, choose "top" or "full"')

    def _existing_level_cols(df, level=1, address='top'):
        newcols = [x[level] for x in list(df.columns)]
        if address == 'top':
            newcols = dict(zip([x[0] for x in df.columns], newcols))
        elif address == 'full':
            newcols = dict(zip(df.columns.levels, newcols))
        return newcols

    def _newcols_generator(dfinternal, level):
        if isinstance(dfinternal.columns, pd.Index) and not isinstance(dfinternal.columns,
                                                                       pd.MultiIndex):  # if only 1D index, must be asking for new column level
            newcolsfn = _new_level_emptycols
        elif len(dfinternal.columns.levels) - 1 < level:  # if asking for new level
            newcolsfn = _new_level_emptycols
        else:  # column labels already exist
            newcolsfn = _existing_level_cols
        return newcolsfn

    dfinternal = df[:]  # shallow copy to prevent changing later df's
    if level == 0:
        raise ValueError("Using level 0 will overwrite main column titles")
    if type(on_cols) != list:
        on_cols = [on_cols]
    if type(on_cols[0]) == tuple:  # if fully addressing with tuples
        address = 'full'
    else:
        address = 'top'

    newcolsfn = _newcols_generator(df, level)  # Either gets _new... or _existing... colnames
    newcols = newcolsfn(dfinternal, level, address=address)

    for col in on_cols:  # Set new values of columns
        if col not in newcols.keys() and col != '.':
            raise KeyError(f'Column ({col}) does not exist in df')
        elif col == '.':
            continue
        newcols[col] = new_col
    if isinstance(dfinternal.columns, pd.Index) and not isinstance(dfinternal.columns, pd.MultiIndex):
        colarray = [list(dfinternal.columns)]
    else:
        colarray = []
        for i in [x for x in range(len(dfinternal.columns.levels)) if x != level]:
            colarray.append([x[i] for x in dfinternal.columns])
    colarray.append(list(newcols.values()))
    dfinternal.columns = pd.MultiIndex.from_arrays(colarray)
    return dfinternal


def add_new_col(df, coladdress):
    dfinternal = df[:]
    coladdress = CU.ensure_list(coladdress)
    dfinternal = _add_col_depth(dfinternal, len(coladdress))  # Adds depth to columns if necessary
    if len(coladdress) == 1:
        dfinternal[coladdress[0]] = ''  # basic add column
        return dfinternal
    else:
        coladdress = coladdress + list(np.repeat('', len(dfinternal.columns.levels) - len(coladdress)))
        dfinternal[tuple(coladdress)] = ''
        return dfinternal


def _add_col_depth(df, depth):
    """Adds depth to columns if necessary (i.e. if trying to add 3 level column to df with only 1 level this
    will add two blank rows first)"""
    dfinternal = df[:]
    if depth > 1:
        if not hasattr(dfinternal.columns, 'levels'):
            levels = 1
            while levels < depth:
                dfinternal = add_col_label(dfinternal, '', '', level=levels)  # add blank levels until
                # depth is great enough
                levels = len(dfinternal.columns.levels)  # Should be multi-indexed now so will have 'levels' attr
    return dfinternal


def get_single_value_pd(df, index, coladdress: tuple):
    assert type(coladdress) == tuple
    ret = df.loc[index, coladdress]
    if isinstance(ret, pd.Series):
        if ret.size != 1:
            raise ValueError(f'Index "{index}", Col "{coladdress}" specifies a series with size "{ret.size}" != 1')
        else:
            ret = ret[0]
    return ret


def get_dtype(df, coladdress: tuple) -> type:
    assert type(coladdress) == tuple
    if len(coladdress) == 1:
        ret = df[coladdress[0]]
    else:
        ret = df[coladdress]
    if isinstance(ret, pd.Series):
        return ret.dtype
    if isinstance(ret, pd.DataFrame):
        if ret.shape[1] != 1:
            raise ValueError(f'Col "{coladdress}" specifies a DataFrame with "{ret.shape[0]}" columns')
        else:
            return ret.dtypes[0]
    else:
        raise UnboundLocalError('Not supposed to get here')


def is_null(df, index, coladdress):
    """Returns True if addressed cell is np.NaN, false otherwise.
    Had to make this to handle getting either a bool or series back when addressing. Set to raise
    error if addressing more than one cell"""
    truth = pd.isna(df.loc[index, coladdress])
    if type(truth) == bool:
        return truth
    elif type(truth) == pd.Series:
        if truth.__len__() > 1:
            raise ValueError("Addressing more than one cell with _is_null")
        return truth.all()  # Returns False if any false, True if all True


def df_backup(func):
    @wraps
    def wrapper(*args, **kwargs):
        return func(*args, **kwargs)

    return wrapper


class inst_dict(MutableMapping):
    """
    Clever dictionary which adds current config name to key when setting values, and add current config name to key when getting values
    """

    def __init__(self, *args, **kwargs):
        self.store = dict()
        self.update(dict(*args, **kwargs))  # use the free update to set keys

    def __getitem__(self, key):
        return self.store[self.__keytransform__(key)]

    def __setitem__(self, key, value):
        self.store[self.__keytransform__(key)] = value

    def __delitem__(self, key):
        del self.store[self.__keytransform__(key)]

    def __iter__(self):
        return iter(self.store)

    def __len__(self):
        return len(self.store)

    def __repr__(self):
        return self.store.__repr__()

    def __str__(self):
        return self.store.__str__()

    def __keytransform__(self, key):
        key = f'{key}_[{cfg.current_config.__name__}]'
        return key
